import openpyxl
import datetime
import calendar
from openpyxl.utils import get_column_letter

from .excel import Excel
import excel2img


# noinspection PyMethodMayBeStatic
class CreateExcel(Excel):
    def __init__(self, values=None):
        super().__init__()

    # noinspection PyAttributeOutsideInit
    def createExcelFile(self):
        self.work = self.getLastInputOf("WORK")
        oid, title, type, folderPath, sheetName, firstCol, Num = self.work

        self.filePath = str(folderPath + f"/{title}.xlsx")

        self.wb = openpyxl.Workbook()
        self.wb.save(self.filePath)
        self.sheet = self.createWorkSheet(self.wb, sheetName, oid)

        self.wb.save(self.filePath)

        self.getPreview(self.filePath, title, self.sheet.title)

    def createWorkSheet(self, wb, sheetName, oid):
        newName = "Sheet"
        if sheetName == "Month Name":
            newName = calendar.month_name[datetime.date.today().month]

        elif sheetName == "Date":
            newName = datetime.date.today().strftime("%d-%m-%y")

        elif sheetName == "WeekDay":
            newName = datetime.date.today().strftime("%A")

        elif sheetName == "Increment":
            newName = oid

        # Check if sheet Already Created by that name
        if newName not in wb.get_sheet_names():
            wb.create_sheet(newName)
            return wb.get_sheet_by_name(newName)
        else:
            name = wb.get_sheet_by_name(newName)
            return name

    def getPreview(self, filePath, title, sheetTitle):
        try:
            excel2img.export_img(filePath, f"static/images/{title}.png", sheetTitle, None)
        except Exception:
            print("Preview cannot update to fix this restart program")

    def addToExcel(self, fields, oid):
        # getting work data and filePath
        work = self.getAllById("WORK", oid)
        oid, title, type, folderPath, sheetName, firstCol, limit = work[0]
        filePath = folderPath + f"/{title}.xlsx"

        # creating connection with excel file
        wb = openpyxl.load_workbook(filePath)
        activeSheet = self.createWorkSheet(wb, sheetName, oid)

        for index, label in enumerate(fields, start=2):
            labelData, labelName = label
            columnLetter = get_column_letter(index)
            column = activeSheet[columnLetter]
            columnLength = 1
            for row in column:
                if row.value and row.value is not None and row.value != '':
                    columnLength += 1

            if firstCol != "None":
                columnL = get_column_letter(index)
                col = activeSheet[columnL]
                colLen = len(col) + 1
                firstCol = self.addFirstColumn(firstCol, activeSheet, colLen, 1, colLen)
                if firstCol is not None:
                    self.insertHistory(title, "First Column", str(firstCol))

            if str(labelData).isnumeric():
                labelData = int(labelData)

            titleIsNotAdded = columnLength < 2
            if titleIsNotAdded:
                activeSheet.cell(row=columnLength, column=index).value = labelName
                columnLength += 1

            cell = activeSheet.cell(row=columnLength, column=index)
            cell.value = labelData
            self.insertHistory(title, labelName, labelData)

        wb.save(filePath)

        self.getPreview(filePath, title, activeSheet.title)

    def addFirstColumn(self, firstCol, wb, columnLen, column, row):
        if firstCol == "Date":
            return self.dateColumn(wb, columnLen, column, row)
        elif firstCol == "Number":
            return self.numberColumn(wb, columnLen, column, row)
        elif firstCol == "formula":
            return self.formulaColumn(wb, columnLen, column, row)

    def dateColumn(self, wb, columnLen, column, row):
        titleCell = wb.cell(row=1, column=column)
        cell = wb.cell(row=row, column=column)
        if columnLen < 2 or titleCell.value != "Date":
            titleCell.value = "Date"
            return "Date"
        else:
            cell.value = datetime.date.today().strftime("%d-%m-%y")
            return datetime.date.today().strftime("%d-%m-%y")

    def numberColumn(self, wb, columnLen, column, row):
        titleCell = wb.cell(row=1, column=column)
        cell = wb.cell(row=row, column=column)
        if columnLen < 2 or titleCell.value != "Date":
            titleCell.value = "Number"
            return "Number"
        else:
            cell.value = row - 1
            return row - 1

    def formulaColumn(self, wb, columnLen, column, row):
        pass

    def replaceLastRows(self, fields, oid):
        # getting work data and filePath
        work = self.getAllById("WORK", oid)
        oid, title, type, folderPath, sheetName, firstCol, limit = work[0]
        filePath = folderPath + f"/{title}.xlsx"

        # creating connection with excel file
        wb = openpyxl.load_workbook(filePath)
        activeSheet = self.createWorkSheet(wb, sheetName, oid)

        for index, label in enumerate(fields, start=1):
            labelData, labelName = label
            columnLetter = get_column_letter(index)
            column = activeSheet[columnLetter]
            columnLength = 1
            for row in column:
                if row.value and row.value is not None and row.value != '':
                    columnLength += 1

            if str(labelData).isnumeric():
                labelData = int(labelData)

            if labelName == "First Column":
                columnLength += 1

            cell = activeSheet.cell(row=columnLength - 1, column=index)
            cell.value = labelData
            self.insertHistory(title, labelName, labelData)

        wb.save(filePath)

        self.getPreview(filePath, title, activeSheet.title)
